import argparse
import copy
import sys
from git import Repo
from .utils import *
from openpyxl import load_workbook
import xmltodict

FILE_ENTRY = {
    'contributor': '',
    'description': '',
    'md5': '',
    'filepath': '',
    'version': 1,
    'published': False,
    'associated_files': [],
    'features': [],
    'notes': ''
}

# ignore files with the following extensions
EXCLUDELIST = ['', '.xls', '.xlsx', '.js', '.html', '.xml', '.dat', '.doc', '.docx', '.txt', '.m4s']

def contribute_files():
    parser = argparse.ArgumentParser(description='Contribute new files to the conformance suite')
    parser.add_argument('-i', '--input', help='Input directory with new files', required=True)
    args = parser.parse_args()

    print('Add (or update) conformance files in "under_consideration" folder.\n')
    contributor = input('Your company name: ')
    input_contrib = input('MPEG document number: ')
    notes_text = ''
    if len(input_contrib) != 0:
        notes_text = f'please refer to {input_contrib} for more information'

    # get rid of the last slash if needed
    if args.input[-1] == '/':
        args.input = args.input[:-1]
    if not os.path.isdir(args.input):
        print('ERROR: input should be a directory')
        sys.exit(-1)
    input_base, input_last_folder = os.path.split(args.input)
    output_root = os.path.join('../../file_features/under_consideration', input_last_folder)

    print(f'Process files from "{input_last_folder}".')

    cnt = 0
    for root, subdirs, files in os.walk(args.input):
        for f in files:
            input_filename, input_extension = os.path.splitext(f)
            if input_extension in EXCLUDELIST:
                continue
            input_path = os.path.join(root, f)
            if ' ' in input_path:
                print(f'ERROR: spaces are not allowed. Remove spaces in "{input_path}"')
                sys.exit(-1)
            
            relative_path = os.path.relpath(root, args.input)
            mpegfs_rel_path = os.path.join('under_consideration', input_last_folder, relative_path, f)
            mpegfs_rel_path = mpegfs_rel_path.replace('./', '')
            output_dir = os.path.join(output_root, relative_path)
            output_filename = input_filename + '.json'
            output_path = os.path.join(output_dir, output_filename)

            print(f'\n* Processing: "{input_path}"')
            md5 = compute_file_md5(input_path)

            file_entry = None
            version = 1
            if os.path.exists(output_path):
                existing_data = read_json(output_path)
                if existing_data['md5'] == md5:
                    print(f'skip duplicate file.')
                    continue
                else:
                    print(f'* new file version detected for {f}')
                    file_entry = existing_data
                    version = file_entry['version'] + 1
                    description = file_entry['description']

            make_dirs_from_path(output_dir)
            if file_entry is None:
                description = input(f'* Provide short description for "{f}": ')
            else:
                new_description = input(f'* Update description for "{f}" (leave blank to keep the previous version): ')
                if len(new_description) > 0:
                    description = new_description
            
            if file_entry is None:
                file_entry = copy.deepcopy(FILE_ENTRY)
            file_entry['contributor'] = contributor.strip()
            file_entry['md5'] = md5
            file_entry['filepath'] = mpegfs_rel_path
            file_entry['description'] = description
            file_entry['notes'] = notes_text
            file_entry['version'] = version
            file_entry['published'] = False
            dump_to_json(output_path, file_entry)
            cnt += 1
    print(f'Processed {cnt} files.')

    # remove part of the filepath (because we operate on our local dir here)
    ignore_prefix = os.path.join('under_consideration', input_last_folder)
    _extract_file_boxes_gpac(output_root, args.input, ignore_prefix)


def update_heif_features():
    parser = argparse.ArgumentParser(description='Update HEIF file features using the excel')
    parser.add_argument('-i', '--inputExcel', help='Input excel sheet', required=True)
    parser.add_argument('-d', '--fileDir', help='Directory with the HEIF json files', required=True)
    args = parser.parse_args()

    wb = load_workbook(args.inputExcel)
    ws = wb.active
    if not ws['A1'].value.lower() == 'file id' \
            or not ws['B1'].value.lower() == 'description' \
            or not ws['C1'].value.lower() == 'input bitstreams':
        print(f'ERROR: worksheet "{args.inputExcel}" is not supported.')
        sys.exit(0)
    # iterate rows
    heif_files = {}
    bitstream_files = {}
    rows = tuple(ws.rows)
    bitstream_id_found = False
    for row in rows:
        if row[0].value is None:
            continue
        elif row[0].value.lower() == 'bitstream id':
            bitstream_id_found = True
            continue
        if bitstream_id_found:
            bitstream_files[row[0].value] = {
                'description': row[1].value.strip()
            }
    for row in rows:
        if row[0].value is None:
            continue
        elif row[0].value.lower() == 'file id':
            continue
        elif row[0].value.lower() == 'bitstream id':
            break
        bitstream_names = [i.strip() for i in row[2].value.split(',')]
        bitstreams = {}
        for bs_name in bitstream_names:
            if bs_name not in bitstream_files:
                print(f'WARNING: bitstream "{bs_name}" is not recognized')
                continue
            bitstreams[bs_name] = bitstream_files[bs_name]

        heif_files[row[0].value] = {
            'description': row[1].value.strip(),
            'bitstreams': bitstreams
        }
    # iterate json files and update them
    for root, subdirs, files in os.walk(args.fileDir):
        for f in files:
            input_filename, input_extension = os.path.splitext(f)
            if not input_extension == '.json':
                continue
            if input_filename not in heif_files:
                print(f'"{input_filename}" has no associated entry in the provided excel sheet')
                continue
            input_path = os.path.join(root, f)
            with open(input_path, 'r') as file:
                data = json.load(file)
                data['description'] = heif_files[input_filename]['description']
                data['notes'] = {
                    'bitstreams': heif_files[input_filename]['bitstreams']
                }

            dump_to_json(input_path, data)


def update_ff_conformance_xls():
    parser = argparse.ArgumentParser(description='Update file features based on ff-conformance.xls processed in '
                                                 'podborski/isobmff-conformance.git repository')
    parser.add_argument('-d', '--fileDir', help='Directory with the file feature json files', required=True)
    args = parser.parse_args()

    if not os.path.exists('temp'):
        Repo.clone_from('https://github.com/podborski/isobmff-conformance.git', 'temp')

    # iterate json files from https://github.com/podborski/isobmff-conformance.git
    podborski_files = {}
    for root, subdirs, files in os.walk('temp'):
        for f in files:
            input_filename, input_extension = os.path.splitext(f)
            if not input_extension == '.json' or '_excel' not in input_filename:
                continue
            input_path = os.path.join(root, f)
            relative_path = os.path.relpath(root, 'temp/data/file_features')
            input_filename = input_filename.replace('_excel', '')
            key = os.path.join(relative_path, os.path.splitext(input_filename)[0])
            podborski_files[key] = input_path

    for root, subdirs, files in os.walk(args.fileDir):
        for f in files:
            input_filename, input_extension = os.path.splitext(f)
            if not input_extension == '.json':
                continue
            input_path = os.path.join(root, f)
            relative_path = os.path.relpath(root, args.fileDir)
            key = os.path.join(relative_path, input_filename)
            if key not in podborski_files:
                print(f'key "{key}" not found in ff-conformance.xls')
                continue
            with open(podborski_files[key], 'r') as file:
                excel_data = json.load(file)
            with open(input_path, 'r') as file:
                file_data = json.load(file)

            file_data['contributor'] = excel_data['company']
            file_data['description'] = excel_data['concept']
            file_data['features'] = excel_data['features']

            dump_to_json(input_path, file_data)


def get_gpac_version():
    process = subprocess.Popen(['MP4Box', '-version'], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    out, err = process.communicate()
    version = err.decode("utf-8").split('\n')[0].strip()
    return version

def _extract_file_boxes_gpac(fileDir, rootDir, ignorePath):
    print(f'Run MP4Box on files in "{rootDir}". Process files as specified in "{fileDir}"')
    mp4box_version = get_gpac_version()

    ret_code = execute_cmd('MP4Box -version')
    if not ret_code == 0:
        print('MP4Box is not installed on your system')
        sys.exit(-1)

    for root, subdirs, files in os.walk(fileDir):
        for filename in files:
            filename_noext, input_extension = os.path.splitext(filename)
            if not input_extension == '.json' or '_gpac' in filename_noext:
                continue
            input_path = os.path.join(root, filename)
            with open(input_path, 'r') as f:
                json_data = json.load(f)
            if 'filepath' not in json_data:
                print(f'Skip {input_path} as no "filepath" key was found.')
                continue
            if ignorePath is None:
                mp4_path = os.path.join(rootDir, json_data['filepath'])
            else:
                relpath = os.path.relpath(json_data['filepath'], ignorePath)
                mp4_path = os.path.join(rootDir, relpath)
            if not os.path.exists(mp4_path):
                print(f'WARNING: mp4 file "{mp4_path}" found in {filename} does not exist!')
                continue
            out_path = os.path.splitext(input_path)[0] + '_gpac.json'
            # pipe MP4Box to stdout and store it in out
            process = subprocess.Popen(['MP4Box', '-diso', '-std', mp4_path], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            out, err = process.communicate()
            try:
                gpac_dict = xmltodict.parse(out)
            except xmltodict.expat.ExpatError:
                print(f'WARNING: Skip "{filename}". Not valid xml output from "{mp4_path}".')
                continue
            gpac_dict['mp4boxVersion'] = mp4box_version
            dump_to_json(out_path, gpac_dict)

def extract_file_boxes_gpac():
    parser = argparse.ArgumentParser(description='Extract box structure using MP4Box')
    parser.add_argument('-d', '--fileDir', help='Directory with the file feature json files', required=True)
    parser.add_argument('-r', '--rootDir', help='Root directory where conformance files are hosted', required=True)
    args = parser.parse_args()

    _extract_file_boxes_gpac(args.fileDir, args.rootDir, None)
