import os
import argparse
import copy
import sys
import json
import subprocess
import shutil
from glob import glob
from git import Repo
from .utils import (
    dump_to_json,
    read_json,
    compute_file_md5,
    make_dirs_from_path,
    execute_cmd,
    file_to_text,
)
from openpyxl import load_workbook
import xmltodict
from functools import cache
import xml.etree.ElementTree as elementTree
from jsonschema import validate
from jsonschema.exceptions import ValidationError

from common import get_mp4ra_boxes

FILE_ENTRY = {
    "contributor": "",
    "description": "",
    "mdms_number": "",
    "rel_filepath": "",
    "associated_files": [],
    "version": 1,
    "md5": "",
    "published": False,
    "features": [],
    "conforms_to": [],
    "notes": "",
    "license": "",
}

# ignore files with the following extensions
EXCLUDELIST = [
    "",
    ".xls",
    ".xlsx",
    ".js",
    ".html",
    ".xml",
    ".dat",
    ".doc",
    ".docx",
    ".txt",
    ".m4s",
]


def contribute_files():
    parser = argparse.ArgumentParser(
        description="Contribute new files to the conformance framework"
    )
    parser.add_argument(
        "-i",
        "--input",
        help="File(s) to contribute. Glob patterns supported.",
        required=True,
    )
    parser.add_argument(
        "-f",
        "--force",
        help="Re-run MP4Box even if the md5's match.",
        action="store_true",
    )
    parser.add_argument(
        "-l",
        "--license",
        help="Path to a .txt file with a license",
    )
    args = parser.parse_args()

    # Load schemas for validation
    with open("../data/schemas/file-metadata.schema.json", "r", encoding="utf-8") as f:
        file_metadata_schema = json.load(f)

    # gather all files we need to process
    files_to_process = glob(args.input, recursive=True)
    print(f"File(s) to process: {len(files_to_process)}")

    contributor_user = input("Contributor company name: ").strip()
    mdms_number_user = input("MPEG document number: ").strip()
    license_str = ""

    if args.license is None:
        print("WARNING: You didn't provide a license for files you want to contribute.")
        response = input(" Are you sure you want to continue? (y - yes): ")
        if not response == "y" and not response == "yes":
            print("Abort")
            sys.exit(1)
    else:
        with open(args.license, "r", encoding="utf-8") as file:
            license_str = file.read().strip()

    cnt = 0
    for f in files_to_process:
        input_root, input_ext = os.path.splitext(f)
        if input_ext in EXCLUDELIST:
            print(f"Skip: {f}")
            continue
        if " " in f:
            print(f'ERROR: spaces are not allowed. Remove spaces in "{f}"')
            sys.exit(-1)
        if "../data/file_features/" not in f:
            print(f"ERROR: file is not inside ../data/file_features/ directory: {f}")
            sys.exit(-1)

        json_path = input_root + ".json"
        rel_filepath = "./" + os.path.splitext(os.path.basename(f))[0] + input_ext
        json_gpac_path = input_root + "_gpac.json"
        json_gpac_ext_path = input_root + "_gpac.ext.json"

        file_md5 = compute_file_md5(f)
        version = 1
        description = ""

        json_data = None
        if os.path.exists(json_path):
            json_data = read_json(json_path)
            validate(instance=json_data, schema=file_metadata_schema)

        # json file found
        if json_data is not None:
            contributor = json_data["contributor"]
            description = json_data["description"]
            mdms_number = json_data["mdms_number"]
            rel_filepath = json_data["rel_filepath"]
            version = json_data["version"]

            # is it a new version of the file?
            if file_md5 == json_data["md5"]:
                if args.force is False:
                    print(f"Skip duplicate file: {f}")
                    continue
            else:
                print(f"New file version detected for {f}")
                version = json_data["version"] + 1

            # new description?
            new_description = input(
                f'* Update description for "{f}" (leave blank to keep the previous version): '
            )
            if len(new_description) > 0:
                description = new_description
        else:
            contributor = contributor_user
            mdms_number = mdms_number_user
            json_data = copy.deepcopy(FILE_ENTRY)
            # try getting the description from the <filename>.txt
            descr_file_text = file_to_text(input_root + ".txt")
            if descr_file_text is None:
                description = input(f'* Provide short description for "{f}": ')
            else:
                description = descr_file_text

        json_data["contributor"] = contributor
        json_data["description"] = description
        json_data["mdms_number"] = mdms_number
        json_data["rel_filepath"] = rel_filepath
        json_data["version"] = version
        json_data["md5"] = file_md5
        json_data["license"] = license_str
        dump_to_json(json_path, json_data)

        # run MP4Box and update if needed
        gpac_dict = _run_mp4box_on_file(f)
        gpac_dict_old = read_json(json_gpac_path)
        if gpac_dict_old is not None:
            if gpac_dict_old["mp4boxVersion"] != gpac_dict["mp4boxVersion"]:
                print(
                    f'WARNING: GPAC file for "{f}" already exists but it has a different MP4Box version. Forcing overwrite!'
                )
                dump_to_json(json_gpac_path, gpac_dict)
        else:
            dump_to_json(json_gpac_path, gpac_dict)

        # Create GPAC extension
        mp4ra_check = "under_consideration" not in os.path.dirname(json_path)
        unknown_boxes = traverse_gpac_dict(gpac_dict, mp4ra_check)
        if len(unknown_boxes) > 0:
            gpac_extension = {
                "mp4boxVersion": gpac_dict["mp4boxVersion"],
                "rel_filepath": rel_filepath,
                "extensions": unknown_boxes,
            }
            if os.path.exists(json_gpac_ext_path):
                print(f'WARNING: "{json_gpac_ext_path}" already exists.')
            else:
                dump_to_json(json_gpac_ext_path, gpac_extension)

        cnt += 1
    print(f"Processed {cnt} files.")


def _isXml(value):
    try:
        elementTree.fromstring(value)
    except elementTree.ParseError:
        return False
    return True


def _run_mp4box_on_file(input_path):
    process = subprocess.Popen(
        [
            "MP4Box",
            "-diso",
            "-std",
            "-logs",
            "all@error",
            "-no-check",
            input_path,
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    out, err = process.communicate()

    try:
        if err and not _isXml(out.decode("utf-8")):
            raise Exception(err)
        gpac_dict = xmltodict.parse(out)
    except xmltodict.expat.ExpatError as ex:
        print(
            f'ERROR: Couldn\'t parse "{input_path}". Not valid xml output from MP4Box.'
        )
        raise ex
    except Exception as ex:
        print(f'ERROR: Couldn\'t parse "{input_path}". MP4Box error.')
        print(err.decode("utf-8"))
        raise ex

    gpac_dict["mp4boxVersion"] = get_gpac_version()
    return gpac_dict


def update_file_features():
    parser = argparse.ArgumentParser(
        description="Try to update file features with latest available MP4Box version"
    )
    parser.add_argument(
        "-i",
        "--input",
        help="Input glob pattern for the files to update",
        type=str,
        default="../data/file_features/**/*.json",
    )
    parser.add_argument(
        "--dry",
        help="Dry run. Do not update the files. Only print the changes.",
        action="store_true",
    )

    args = parser.parse_args()

    # Get all metadata files
    metadata_files = glob(args.input, recursive=True)

    # Filter out gpac files
    metadata_files = list(filter(lambda x: "_gpac" not in x, metadata_files))

    # Traverse all files
    updated_files = 0
    for metadata_file in metadata_files:
        with open(metadata_file, "r") as f:
            metadata = json.load(f)

        # Path to the actual file
        actual_file_path = os.path.join(
            os.path.dirname(metadata_file), metadata["rel_filepath"]
        )

        # If the file is ZIPed, decompress it
        if os.path.splitext(actual_file_path)[1] == ".zip":
            print(f'Unzipping "{actual_file_path}"')
            ret = execute_cmd(
                f"unzip -o {actual_file_path} -d {os.path.dirname(actual_file_path)}"
            )
            if ret != 0:
                print(f'ERROR: Failed to unzip "{actual_file_path}"')
                continue
            actual_file_path = os.path.splitext(actual_file_path)[0]

        print(f'Processing "{actual_file_path}"')
        try:
            gpac_dict = _run_mp4box_on_file(actual_file_path)
        except Exception:
            continue

        # If the file was ZIPed, remove the decompressed files
        if os.path.splitext(metadata["rel_filepath"])[1] == ".zip":
            print(f'Removing decompressed "{actual_file_path}"')
            os.remove(actual_file_path)

        # Save the new GPAC file
        gpac_file_path = os.path.splitext(metadata_file)[0] + "_gpac.json"
        if not args.dry:
            dump_to_json(gpac_file_path, gpac_dict)
            updated_files += 1

        # Create extension file if needed
        gpac_extension_path = gpac_file_path.replace("_gpac", "_gpac.ext")
        mp4ra_check = "under_consideration" not in os.path.dirname(metadata_file)
        unknown_boxes = traverse_gpac_dict(gpac_dict, mp4ra_check)

        # Remove GPAC extension if there are no unknown boxes anymore
        if len(unknown_boxes) == 0 and os.path.exists(gpac_extension_path):
            os.remove(gpac_extension_path)

        # Create GPAC extension
        if len(unknown_boxes) > 0:
            gpac_extension = {
                "mp4boxVersion": gpac_dict["mp4boxVersion"],
                "rel_filepath": metadata["rel_filepath"],
                "extensions": unknown_boxes,
            }
            if not args.dry:
                dump_to_json(gpac_extension_path, gpac_extension)

    print(
        f"{args.dry and '[DRY] ' or ''}Updated {updated_files} ({len(metadata_files)}) files."
    )


def update_heif_features():
    parser = argparse.ArgumentParser(
        description="Update HEIF file features using the excel"
    )
    parser.add_argument("-i", "--inputExcel", help="Input excel sheet", required=True)
    parser.add_argument(
        "-d", "--fileDir", help="Directory with the HEIF json files", required=True
    )
    args = parser.parse_args()

    wb = load_workbook(args.inputExcel)
    ws = wb.active
    if (
        not ws["A1"].value.lower() == "file id"
        or not ws["B1"].value.lower() == "description"
        or not ws["C1"].value.lower() == "input bitstreams"
    ):
        print(f'ERROR: worksheet "{args.inputExcel}" is not supported.')
        sys.exit(0)
    # iterate rows
    heif_files = {}
    bitstream_files = {}
    rows = tuple(ws.rows)
    bitstream_id_found = False
    for row in rows:
        if row[0].value is None:
            continue
        elif row[0].value.lower() == "bitstream id":
            bitstream_id_found = True
            continue
        if bitstream_id_found:
            bitstream_files[row[0].value] = {"description": row[1].value.strip()}
    for row in rows:
        if row[0].value is None:
            continue
        elif row[0].value.lower() == "file id":
            continue
        elif row[0].value.lower() == "bitstream id":
            break
        bitstream_names = [i.strip() for i in row[2].value.split(",")]
        bitstreams = {}
        for bs_name in bitstream_names:
            if bs_name not in bitstream_files:
                print(f'WARNING: bitstream "{bs_name}" is not recognized')
                continue
            bitstreams[bs_name] = bitstream_files[bs_name]

        heif_files[row[0].value] = {
            "description": row[1].value.strip(),
            "bitstreams": bitstreams,
        }
    # iterate json files and update them
    for root, subdirs, files in os.walk(args.fileDir):
        for f in files:
            input_filename, input_extension = os.path.splitext(f)
            if not input_extension == ".json":
                continue
            if input_filename not in heif_files:
                print(
                    f'"{input_filename}" has no associated entry in the provided excel sheet'
                )
                continue
            input_path = os.path.join(root, f)
            with open(input_path, "r") as file:
                data = json.load(file)
                data["description"] = heif_files[input_filename]["description"]
                data["notes"] = {"bitstreams": heif_files[input_filename]["bitstreams"]}

            dump_to_json(input_path, data)


def update_ff_conformance_xls():
    parser = argparse.ArgumentParser(
        description="Update file features based on ff-conformance.xls processed in "
        "podborski/isobmff-conformance.git repository"
    )
    parser.add_argument(
        "-d",
        "--fileDir",
        help="Directory with the file feature json files",
        required=True,
    )
    args = parser.parse_args()

    if not os.path.exists("temp"):
        Repo.clone_from("https://github.com/podborski/isobmff-conformance.git", "temp")

    # iterate json files from https://github.com/podborski/isobmff-conformance.git
    podborski_files = {}
    for root, subdirs, files in os.walk("temp"):
        for f in files:
            input_filename, input_extension = os.path.splitext(f)
            if not input_extension == ".json" or "_excel" not in input_filename:
                continue
            input_path = os.path.join(root, f)
            relative_path = os.path.relpath(root, "temp/data/file_features")
            input_filename = input_filename.replace("_excel", "")
            key = os.path.join(relative_path, os.path.splitext(input_filename)[0])
            podborski_files[key] = input_path

    for root, subdirs, files in os.walk(args.fileDir):
        for f in files:
            input_filename, input_extension = os.path.splitext(f)
            if not input_extension == ".json":
                continue
            input_path = os.path.join(root, f)
            relative_path = os.path.relpath(root, args.fileDir)
            key = os.path.join(relative_path, input_filename)
            if key not in podborski_files:
                print(f'key "{key}" not found in ff-conformance.xls')
                continue
            with open(podborski_files[key], "r") as file:
                excel_data = json.load(file)
            with open(input_path, "r") as file:
                file_data = json.load(file)

            file_data["contributor"] = excel_data["company"]
            file_data["description"] = excel_data["concept"]
            file_data["features"] = excel_data["features"]

            dump_to_json(input_path, file_data)


@cache
def get_gpac_version():
    process = subprocess.Popen(
        ["MP4Box", "-version"], stdout=subprocess.PIPE, stderr=subprocess.PIPE
    )
    out, err = process.communicate()
    version = err.decode("utf-8").split("\n")[0].strip()
    return version


def _extract_file_boxes_gpac(fileDir):
    print(f'Run MP4Box on files in "{fileDir}"')
    mp4box_version = get_gpac_version()

    ret_code = execute_cmd("MP4Box -version")
    if not ret_code == 0:
        print("MP4Box is not installed on your system")
        sys.exit(-1)

    for root, subdirs, files in os.walk(fileDir):
        for filename in files:
            filename_noext, input_extension = os.path.splitext(filename)
            if not input_extension == ".json" or "_gpac" in filename_noext:
                print(f"Skip {filename}")
                continue
            input_path = os.path.join(root, filename)
            json_data = read_json(input_path)

            if "rel_filepath" not in json_data:
                print(f'Skip {input_path} as no "rel_filepath" key was found.')
                continue

            mp4_path = os.path.join(root, json_data["rel_filepath"])
            if not os.path.exists(mp4_path):
                print(f'WARNING: mp4 file "{mp4_path}" not found!')
                continue
            out_path = os.path.splitext(input_path)[0] + "_gpac.json"
            # pipe MP4Box to stdout and store it in out
            try:
                gpac_dict = _run_mp4box_on_file(actual_file_path)
            except Exception:
                continue
            dump_to_json(out_path, gpac_dict)


def traverse_gpac_dict(gpac_dict, mp4ra_check=True):
    """Traverse the GPAC dict and return a list of unknown boxes with their location"""
    unknown_boxes = []

    def add(value, path):
        nonlocal unknown_boxes
        unknown_boxes.append(
            {
                "location": ".".join(path),
                "box": {
                    **value,
                    "@Container": path[-1],
                },
            }
        )

    def crawl(hierarchy, path=[]):
        # If an object has @Type then follow it. record the path
        for key, value in hierarchy.items():
            if isinstance(value, dict):
                if key == "UnknownBox":
                    add(value, path)
                    continue

                if (
                    mp4ra_check
                    and "@Type" in value
                    and value["@Type"] not in get_mp4ra_boxes()
                ):
                    continue

                if "@Type" in value:
                    crawl(value, path + [value["@Type"]])
            elif isinstance(value, list):
                for item in value:
                    crawl({key: item}, path)

    # Special case for gpac files. Root is not a box
    crawl(gpac_dict["IsoMediaFile"], path=["file"])
    return unknown_boxes


def _create_gpac_extension(fileDir):
    print(f'Creating GPAC extensions for files in "{fileDir}"')
    mp4box_version = get_gpac_version()

    ret_code = execute_cmd("MP4Box -version")
    if not ret_code == 0:
        print("MP4Box is not installed on your system")
        sys.exit(-1)

    for root, subdirs, files in os.walk(fileDir):
        for filename in files:
            filename_noext, input_extension = os.path.splitext(filename)

            if "gpac.json" not in filename:
                continue

            # Load non-gpac metadata file
            metadata_path = os.path.join(
                root, filename_noext.replace("_gpac", "") + ".json"
            )
            with open(metadata_path, "r") as f:
                metadata = json.load(f)

            # Go through the GPAC file and figure out unknown boxes
            input_path = os.path.join(root, filename)
            with open(input_path, "r") as f:
                json_data = json.load(f)

            mp4ra_check = "under_consideration" not in root
            unknown_boxes = traverse_gpac_dict(json_data, mp4ra_check)
            if len(unknown_boxes) == 0:
                continue

            gpac_extension = {
                "mp4boxVersion": mp4box_version,
                "rel_filepath": metadata["rel_filepath"],
                "extensions": unknown_boxes,
            }

            # Dump to json
            out_path = os.path.join(root, filename_noext + ".ext.json")
            dump_to_json(out_path, gpac_extension)


def extract_file_features():
    parser = argparse.ArgumentParser(description="Extract box structure using MP4Box")
    parser.add_argument(
        "-i",
        "--input",
        help="The source file to create MP4Box output from",
    )
    parser.add_argument(
        "-d",
        "--dir-input",
        help="The source files to create MP4Box output from. Use glob patterns.",
    )
    parser.add_argument(
        "-f",
        "--overwrite",
        help="Overwrite existing file features",
        action="store_true",
    )
    parser.add_argument(
        "-m",
        "--metadata",
        help="Create empty metadata file for the file",
        action="store_true",
    )
    args = parser.parse_args()

    if args.input is None and args.dir_input is None:
        print("ERROR: You must provide either --input or --dir-input")
        sys.exit(-1)

    if args.input is not None and args.dir_input is not None:
        print("ERROR: You can only provide either --input or --dir-input")
        sys.exit(-1)

    if args.input is not None:
        _extract_file_features(args)

    if args.dir_input is not None:
        files = glob(args.dir_input, recursive=True)
        for f in files:
            args.input = f
            _extract_file_features(args, exit_on_error=False)


def _extract_file_features(args, exit_on_error=True):
    # During CI, we might try to process an associated file.
    # If that's the case, search if any metadata mentions the file
    with open("../data/schemas/file-metadata.schema.json", "r", encoding="utf-8") as f:
        file_metadata_schema = json.load(f)

    metadata_files = glob("../data/file_features/**/*.json", recursive=True)
    for mf in metadata_files:
        if "_gpac" in mf:
            continue

        with open(mf, "r", encoding="utf-8") as f:
            metadata = json.load(f)

        # The associated file might be a JSON, check if it satisfies the schema
        try:
            validate(instance=metadata, schema=file_metadata_schema)
        except ValidationError:
            continue

        real_input_path = os.path.realpath(args.input)
        for af in metadata["associated_files"]:
            real_af_path = os.path.realpath(os.path.join(os.path.dirname(mf), af))
            if real_af_path == real_input_path:
                print(
                    f'WARNING: "{args.input}" is already mentioned in "{mf}". Skipping.'
                )
                if exit_on_error:
                    sys.exit(0)
                else:
                    return

    # File paths
    input_path = args.input
    rel_input_path = f"./{os.path.basename(input_path)}"
    metadata_path = os.path.splitext(input_path)[0] + ".json"
    gpac_path = os.path.splitext(input_path)[0] + "_gpac.json"
    gpac_extension_path = os.path.splitext(gpac_path)[0] + ".ext.json"

    # Check if metadata file exists and if it does, check if it has the same md5 hash
    if os.path.exists(metadata_path):
        md5_gt = read_json(metadata_path)["md5"]
        md5_ref = compute_file_md5(input_path)
        if md5_gt != md5_ref:
            print(
                f'WARNING: Metadata for "{input_path}" already exists but it has a different md5 hash. Forcing overwrite!'
            )
            args.overwrite = True

    # Create metadata file
    if args.metadata:
        if not args.overwrite and os.path.exists(metadata_path):
            print(
                f'WARNING: "{metadata_path}" already exists. Use --overwrite to overwrite.'
            )
        else:
            # If we already have the metadata file read it and update it
            if os.path.exists(metadata_path):
                with open(metadata_path, "r", encoding="utf-8") as f:
                    metadata = json.load(f)
            else:
                metadata = copy.deepcopy(FILE_ENTRY)

            with open(metadata_path, "w", encoding="utf-8") as f:
                # copy the file entry
                metadata["rel_filepath"] = rel_input_path
                metadata["md5"] = compute_file_md5(input_path)
                json.dump(metadata, f, indent=2)

    # Run MP4Box
    try:
        gpac_dict = _run_mp4box_on_file(input_path)
    except Exception:
        if exit_on_error:
            sys.exit(0)  # delegate error handling to CI
        else:
            return

    # If we already have the GPAC file and the contents have been changed, update it
    if os.path.exists(gpac_path):
        with open(gpac_path, "r", encoding="utf-8") as f:
            gpac_dict_gt = json.load(f)

        if gpac_dict_gt["IsoMediaFile"] != gpac_dict["IsoMediaFile"]:
            print(
                f'WARNING: GPAC file for "{input_path}" already exists but the contents have been changed. Forcing overwrite!'
            )
            args.overwrite = True

    if not args.overwrite and os.path.exists(gpac_path):
        print(f'WARNING: "{gpac_path}" already exists. Use --overwrite to overwrite.')
    else:
        dump_to_json(gpac_path, gpac_dict)

    # Create GPAC extension
    mp4ra_check = "under_consideration" not in os.path.dirname(metadata_path)
    unknown_boxes = traverse_gpac_dict(gpac_dict, mp4ra_check)
    if len(unknown_boxes) == 0:
        if exit_on_error:
            sys.exit(0)
        else:
            return

    gpac_extension = {
        "mp4boxVersion": gpac_dict["mp4boxVersion"],
        "rel_filepath": rel_input_path,
        "extensions": unknown_boxes,
    }

    if not args.overwrite and os.path.exists(gpac_extension_path):
        print(
            f'WARNING: "{gpac_extension_path}" already exists. Use --overwrite to overwrite.'
        )
    else:
        dump_to_json(gpac_extension_path, gpac_extension)
